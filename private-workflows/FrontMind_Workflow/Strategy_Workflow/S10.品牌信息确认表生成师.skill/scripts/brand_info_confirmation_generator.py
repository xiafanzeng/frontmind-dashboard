#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
S10 品牌信息确认表生成器 (Brand Info Confirmation Sheet Generator)
================================================================
用途：生成一份《品牌信息确认表.xlsx》，作为策略层 S10 的客户确认产物。
该 xlsx 含两个子表：

  子表1「品牌信息确认」：完整复刻《品牌信息收集表》的全部品牌信息区段与条目
      （品牌基础信息 → 补充材料清单，共 13 个区段），中间设「企业填写 / 修改」
      高亮列供企业校正。删除原收集表末尾「提出的问题」与「AI 可见性监测准备」
      区段（后者归入子表2）。

  子表2「监控问题与应答逻辑确认」：参考《应答逻辑确认表》，以「监控问题 — 应答
      逻辑（预填）」为骨架，中间插入「企业想修改」列，右侧保留「参考资料」。
      （已按要求删除「应答逻辑备注」列。）

两种运行模式：
  1) 模板模式（默认，--mode template 或不带 --work-dir/--response-logic）：
     输出纯空白模板，不预填任何企业名 / 品牌信息，品牌名占位为「【品牌名】」，
     预填列写统一占位提示。适合作为标准模板下发。
  2) 整合模式（--mode fill，需 --work-dir 与 --response-logic）：
     把策略层 S1-S9 成果与企业已填写的《应答逻辑确认表》整合预填进表中。

配色严格遵循《品牌信息收集表》紫色风格：
  主标题 #4A154B / 区段标题 #4A154B / 表头 #6B3FA0 / 交替行 #FFFFFF · #F9F7FC
  企业填写列 #FFFDE7（浅黄）+ 紫字 #4A154B / 必填红字 #D32F2F

输入：
  --brand           品牌名（模板模式可省略，默认「【品牌名】」）
  --mode            template（纯模板，默认）| fill（整合预填）
  --work-dir        策略层成果工作目录（含 S1-S9 源文件）；fill 模式必填
  --response-logic  企业已填写的《应答逻辑确认表》xlsx 路径；fill 模式必填
  --out             输出 xlsx 路径

用法：
  # 生成纯空白模板
  python3 brand_info_confirmation_generator.py --out "./品牌信息确认表_模板.xlsx"

  # 整合预填某品牌
  python3 brand_info_confirmation_generator.py \
      --mode fill --brand "德马克" \
      --work-dir "./strategy_pack_workspace" \
      --response-logic "./德马克_应答逻辑确认表_现场讨论.xlsx" \
      --out "./S10_德马克_品牌信息确认表.xlsx"
"""

import argparse
import glob
import json
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 配色常量（取自《品牌信息收集表》精确色值）
# ---------------------------------------------------------------------------
C_TITLE_BG = "4A154B"       # 主标题 / 区段标题 深紫
C_SUBTITLE_FONT = "D1C4E9"  # 副标题浅紫字
C_HEADER_BG = "6B3FA0"      # 表头 中紫
C_WHITE = "FFFFFF"
C_ROW_ALT = "F9F7FC"        # 交替行 极浅紫
C_FILL_INPUT = "FFFDE7"     # 企业填写 / 修改列 浅黄
C_FONT_INPUT = "4A154B"     # 企业填写列 紫字
C_FONT_REQUIRED = "D32F2F"  # [必填] 红字
C_FONT_HIGH = "6B3FA0"      # [高优] 紫字
C_FONT_PREFILL = "424242"   # 预填内容 深灰
C_FONT_NOTE = "757575"      # 备注 浅灰
C_FONT_ITEM = "1A1A1A"      # 条目名称 近黑
C_FONT_IDX = "9E9E9E"       # 序号灰

# 模板模式下「已有内容（预填）」列的统一占位
PREFILL_PLACEHOLDER = "（由策略层 S1-S9 自动填充）"

THIN = Side(style="thin", color="E0D6EC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def F(color, size=10, bold=False):
    return Font(name="微软雅黑", size=size, bold=bold, color=color)


def fill(color):
    return PatternFill(fill_type="solid", fgColor=color)


WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


# ---------------------------------------------------------------------------
# 工具：在 work-dir 里按通配符找文件并读 JSON
# ---------------------------------------------------------------------------
def find_file(work_dir, *patterns):
    if not work_dir:
        return None
    for p in patterns:
        hits = glob.glob(os.path.join(work_dir, "**", p), recursive=True)
        if hits:
            return hits[0]
    return None


def load_json(path):
    if not path or not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"[warn] 读取 JSON 失败 {path}: {e}", file=sys.stderr)
        return {}


# ---------------------------------------------------------------------------
# 通用：写主标题区（占据所有列）
# ---------------------------------------------------------------------------
def write_title_block(ws, ncols, title, subtitle):
    last_col = get_column_letter(ncols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"].fill = fill(C_TITLE_BG)
    ws.row_dimensions[1].height = 8
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = title
    c.fill = fill(C_TITLE_BG)
    c.font = F(C_WHITE, 20, True)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[2].height = 42
    ws.merge_cells(f"A3:{last_col}3")
    c = ws["A3"]
    c.value = subtitle
    c.fill = fill(C_TITLE_BG)
    c.font = F(C_SUBTITLE_FONT, 10, False)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 6


def write_header_row(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row, i)
        c.value = h
        c.fill = fill(C_HEADER_BG)
        c.font = F(C_WHITE, 10, True)
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 30


def write_section_row(ws, row, ncols, text):
    last_col = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row, 1)
    c.value = text
    c.fill = fill(C_TITLE_BG)
    c.font = F(C_WHITE, 13, True)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 30


# ===========================================================================
# 子表1 结构定义：完整复刻《品牌信息收集表》全部品牌信息区段
# 每个条目 = (条目名称, 填写要求标记, 备注说明)
# 预填内容在 fill 模式下由 S1-S9 注入；template 模式统一用占位提示。
# ===========================================================================
SHEET1_SECTIONS = [
    ("一、品牌基础信息", [
        ("企业全称", "[必填]", "企业工商注册的完整法定名称，用于品牌知识库的基础标识和所有正式文档引用"),
        ("品牌简称", "[必填]", "用于 AI 搜索结果展示、社交媒体传播和日常品牌提及的简短名称"),
        ("所属行业", "[必填]", "明确企业所在细分行业，影响后续营销图谱中的行业关键词和竞争维度分析"),
        ("企业官网", "[高优]", "官方网站地址，用于信息交叉验证和 SEO 基建分析"),
        ("联系方式", "[必填]", "电话、邮箱、微信公众号等，用于品牌知识库中的联系信息展示"),
    ]),
    ("二、品牌定位与价值主张", [
        ("一句话品牌定位", "[必填]", "用一句话概括品牌的核心定位，将直接用于品牌主叙事提炼和 AI 搜索结果中的品牌描述"),
        ("核心价值主张", "[高优]", "用 1-3 句话描述品牌为客户创造的核心价值，将支撑品牌主叙事和内容分类的词根提炼"),
        ("差异化优势", "[高优]", "与竞品相比最突出的 1-3 个差异化优势，用于建立竞争维度和内容差异化策略"),
    ]),
    ("三、产品 / 服务矩阵", [
        ("主营产品线 1", "[必填]", "请确认或修改该产品线的描述、适用场景、关键参数、目标客户、是否重点推广"),
        ("主营产品线 2", "[必填]", "请确认或修改该产品线的描述、适用场景、关键参数、目标客户、是否重点推广"),
        ("主营产品线 3", "[必填]", "请确认或修改该产品线的描述、适用场景、关键参数、目标客户、是否重点推广"),
        ("主营产品线 4", "[必填]", "请确认或修改该产品线的描述、适用场景、关键参数、目标客户、是否重点推广"),
        ("配件 / 耗材 / 零部件", "[必填]", "请确认配件、耗材、零部件等长尾产品线及其复购能力"),
        ("其他产品 / 服务", "[可选]", "如有上述未列出的产品、服务、配件、维修或集成能力，请在企业填写列补充"),
        ("当前重点推广方向", "[必填]", "当前最希望推广的产品线、区域、渠道或客户类型；将影响后续策略优先级"),
    ]),
    ("四、核心优势", [
        ("产品线覆盖广度", "[必填]", "请确认产品覆盖范围是否准确，并补充证据"),
        ("服务链与售后支持", "[必填]", "请确认选型、安装、培训、远程诊断、备件、售后响应等服务链条的真实可公开口径"),
        ("配件耗材与维修长尾", "[必填]", "请补充可复购配件 / 耗材清单、兼容型号、库存与交期口径"),
        ("跨境内容与多平台触达", "[必填]", "请确认官网、阿里国际站、中国制造网、YouTube、LinkedIn、TikTok 等渠道是否为当前重点"),
        ("中小客户与经销商适配性", "[必填]", "请补充面向中小客户、经销商 / 代理商的政策、培训、素材和利润空间口径"),
        ("证据透明化与资料完整度", "[必填]", "请补充可公开引用的证书、案例、产品手册、检测 / 验收标准和客户反馈"),
        ("其他优势", "[可选]", "其他优势请补充，建议以可验证证据支持"),
    ]),
    ("五、目标客群", [
        ("客户类型（B / C 端）", "[必填]", "明确客户是企业客户(B端)还是个人消费者(C端)，影响用户画像研究和内容策略方向"),
        ("主要行业 / 人群分布", "[必填]", "客户所在的主要行业或人群类型，用于用户画像与需求问题拆解"),
        ("客户规模分布", "[必填]", "客户的规模特征，影响内容调性和传播渠道选择"),
        ("地域分布", "[必填]", "客户的地理分布，用于地域化传播策略和本地化内容规划"),
    ]),
    ("六、资质认证", [
        ("ISO9001 / 质量体系认证", "[高优]", "请确认认证名称、编号、范围和证书扫描件；不得把未核实证书作为确定宣传点"),
        ("CE / SGS / 环境或出口相关认证", "[高优]", "请确认 CE、SGS、环境管理或其他出口合规文件的实际持有情况"),
        ("证书编号、有效期、适用型号", "[高优]", "证书扫描件、编号、颁发机构、有效期、适用产品型号、出口文件模板"),
        ("专利、软著与知识产权", "[建议]", "请补充专利、软著、技术证书、测试报告等可公开证据"),
    ]),
    ("七、客户案例", [
        ("案例 1 · 客户名称 / 行业", "[高优]", "客户名称可匿名；请至少填写行业、国家 / 地区、客户类型"),
        ("案例 1 · 需求痛点", "[高优]", "合作前痛点，如效率、质量、成本、售后、备件、交付周期等"),
        ("案例 1 · 解决方案", "[高优]", "设备 / 服务配置、培训、售后或集成方案"),
        ("案例 1 · 合作成果", "[高优]", "合作后的效率、质量、成本、回本周期、复购、转介绍等可量化结果"),
        ("案例 2 · 客户名称 / 行业", "[高优]", "可匿名；请填写国家 / 地区、渠道类型"),
        ("案例 2 · 需求痛点", "[高优]", "合作前痛点"),
        ("案例 2 · 解决方案", "[高优]", "供货 / 服务、培训、售后、素材支持与合作政策"),
        ("案例 2 · 合作成果", "[高优]", "订单、询盘增长、市场覆盖、复购或区域突破等结果"),
        ("更多案例", "[高优]", "强烈建议补充可公开行业、应用场景、配置与结果数据；客户名称可匿名处理"),
    ]),
    ("八、量化经营数据", [
        ("成立年份", "[高优]", "企业成立的具体年份，用于品牌故事与长期可信度建设"),
        ("员工人数", "[高优]", "企业当前员工总数，用于展示企业规模和团队实力"),
        ("年营收规模", "[高优]", "年度营收规模（可用区间），用于增强品牌权威性表达"),
        ("年产能", "[高优]", "年度产能数据，如不适用可注明"),
        ("累计服务客户数", "[高优]", "累计服务的客户总数，用于品牌成果展示"),
        ("典型成果数据", "[高优]", "最具代表性的业务成果数据"),
        ("服务年限", "[高优]", "品牌在该领域的服务年限，用于品牌故事叙事；注意避免被误解为虚假宣传"),
        ("累计项目 / 订单 / 装机量", "[高优]", "累计订单、装机量、典型应用数量等，用于展示服务经验深度"),
        ("转化率 / 复购率 / 售后响应率", "[高优]", "询盘转化率、复购率、售后响应达成率等，用于内容可信度建设"),
        ("重点经销 / 合作伙伴数", "[高优]", "经销商、代理伙伴、平台合作伙伴或重点区域服务资源数量"),
    ]),
    ("九、竞争格局", [
        ("主要竞品 1", "[建议]", "请列出主要竞品名称、竞品优势和我方差异化，便于精准识别关键词竞争与内容空位"),
        ("主要竞品 2", "[建议]", "同上"),
        ("主要竞品 3", "[建议]", "同上"),
        ("更多竞品 / 渠道品牌", "[建议]", "建议补充具体竞品品牌、区域渠道商或客户常比较的供应商"),
        ("客户决策因素排序", "[建议]", "客户选择服务商时最看重的因素排序，用于校准竞争维度和内容优先级"),
    ]),
    ("十、服务区域", [
        ("总部 / 核心区域", "[建议]", "企业总部所在地及核心服务区域"),
        ("重点覆盖区域", "[建议]", "企业重点覆盖的区域或城市"),
        ("全国 / 海外覆盖", "[建议]", "全国或海外的服务覆盖范围"),
    ]),
    ("十一、品牌历史与里程碑", [
        ("创立背景", "[可选]", "企业创立的背景故事，用于品牌故事素材和历史叙事"),
        ("关键里程碑", "[可选]", "企业发展中的关键事件（成立、重要合作、分支设立、重大成果等）"),
        ("未来愿景", "[可选]", "企业的未来发展愿景和战略方向"),
    ]),
    ("十二、备注与想法", [
        ("当前营销困惑或痛点", "[可选]", "记录当前营销推广中的困难或挑战，帮助后续策略更有针对性"),
        ("希望突出的品牌形象 / 调性", "[可选]", "希望在对外传播中重点强调的品牌形象标签，直接影响内容策略调性"),
        ("对目标客群的主观判断", "[可选]", "基于实际经验对客户群体的观察和判断"),
        ("近期计划", "[可选]", "近期业务计划（新产品、市场扩张、品牌升级等），影响营销策略时效性"),
        ("对竞品的观察或担忧", "[可选]", "对竞争对手的主观观察和担忧，帮助完善竞争分析"),
        ("其他补充信息", "[可选]", "任何您觉得有价值但上述未涵盖的信息，均可自由填写"),
    ]),
    ("十三、补充材料清单", [
        ("企业 BP / 宣传册", "[建议]", "如有企业 BP、英文宣传册、经销商手册或展会资料，请提供"),
        ("产品手册 / 目录", "[建议]", "如有产品目录、型号手册、参数表、安装 / 维护手册，请提供"),
        ("官网链接", "[建议]", "企业官方网站链接"),
        ("客户评价 / 感谢信", "[建议]", "如有客户评价、感谢信、验收报告、合作反馈、代理商反馈，请提供"),
        ("行业报告 / 白皮书", "[建议]", "如有行业报告、技术白皮书、应用指南或展会资料，请提供"),
        ("团队成员介绍资料", "[建议]", "如有核心团队、研发、售后、生产、外贸团队的介绍资料，请提供"),
        ("其他资料", "[建议]", "任何其他有助于品牌知识库构建的资料（Logo 源文件、VI 手册、高清图、证书扫描件等）"),
    ]),
]


# ---------------------------------------------------------------------------
# fill 模式：从 S 节点 JSON 中尽力取值；template 模式不调用，返回占位
# 这里采用「宽松抽取」：能取到就填，取不到一律留占位，绝不写入其他企业信息。
# ---------------------------------------------------------------------------
def deep_get(d, *keys, default=""):
    cur = d
    for k in keys:
        if isinstance(cur, dict) and k in cur:
            cur = cur[k]
        else:
            return default
    return cur if cur not in (None, "") else default


def _join_list(v):
    if isinstance(v, list):
        out = []
        for x in v:
            if isinstance(x, dict):
                out.append(x.get("name") or x.get("title") or x.get("text")
                           or json.dumps(x, ensure_ascii=False))
            else:
                out.append(str(x))
        return "；".join(out)
    return str(v) if v else ""


def build_prefill_map(s1, s4, s6):
    """fill 模式下，按条目名称尽力映射 S1/S4/S6 预填值。取不到返回空串。"""
    facts = s1.get("brand_facts", s1) if isinstance(s1, dict) else {}

    def g1(*keys):
        return deep_get(s1, *keys) or deep_get(facts, *keys)

    m = {
        "企业全称": g1("legal_name") or g1("company_name"),
        "品牌简称": g1("brand_name") or g1("brand_short"),
        "所属行业": g1("industry") or g1("category"),
        "企业官网": g1("website") or g1("official_site"),
        "联系方式": " / ".join(x for x in [g1("phone"), g1("email")] if x),
        "一句话品牌定位": deep_get(s4, "positioning_statement") or deep_get(s4, "statement"),
        "核心价值主张": _join_list(deep_get(s4, "value_propositions", default=[]))
        or deep_get(s4, "value_proposition"),
        "差异化优势": _join_list(deep_get(s4, "differentiators", default=[])
                            or deep_get(s4, "differentiation_matrix", default=[])),
        "成立年份": g1("founded_year") or g1("established"),
        "总部 / 核心区域": g1("address") or g1("headquarters"),
        "希望突出的品牌形象 / 调性": _join_list(deep_get(s6, "tone", default=[])
                                     or deep_get(s6, "brand_tone", default=[])),
    }
    return {k: v for k, v in m.items() if v}


# ===========================================================================
# 子表1：品牌信息确认
# ===========================================================================
def build_sheet1(wb, brand, prefill_map, template_mode):
    ws = wb.active
    ws.title = "品牌信息确认"
    headers = ["#", "条目名称", "填写要求", "已有内容（预填）",
               "✏️ 企业填写 / 修改", "备注说明"]
    ncols = len(headers)
    write_title_block(
        ws, ncols,
        f"{brand}  品牌信息确认表",
        "FrontMind 品牌优化工作流  ·  S10 品牌信息确认（S1-S9 整合总结）   "
        "紫色「已有内容」列为自动预填，请在黄色「企业填写 / 修改」列确认、修正或补充",
    )
    write_header_row(ws, 5, headers)

    row = 6
    alt = False
    idx = 0
    for sec_title, items in SHEET1_SECTIONS:
        write_section_row(ws, row, ncols, sec_title)
        row += 1
        for name, flag, note in items:
            idx += 1
            bg = C_ROW_ALT if alt else C_WHITE
            alt = not alt
            if template_mode:
                prefill = PREFILL_PLACEHOLDER
            else:
                prefill = prefill_map.get(name) or PREFILL_PLACEHOLDER

            c = ws.cell(row, 1); c.value = idx; c.fill = fill(bg)
            c.font = F(C_FONT_IDX, 9); c.alignment = CENTER; c.border = BORDER
            c = ws.cell(row, 2); c.value = name; c.fill = fill(bg)
            c.font = F(C_FONT_ITEM, 10, True); c.alignment = WRAP; c.border = BORDER
            c = ws.cell(row, 3); c.value = flag; c.fill = fill(bg)
            fc = (C_FONT_REQUIRED if flag == "[必填]"
                  else C_FONT_HIGH if flag == "[高优]"
                  else C_FONT_NOTE)
            c.font = F(fc, 9, bool(flag)); c.alignment = CENTER; c.border = BORDER
            c = ws.cell(row, 4); c.value = prefill; c.fill = fill(bg)
            c.font = F(C_FONT_PREFILL, 10); c.alignment = WRAP; c.border = BORDER
            c = ws.cell(row, 5); c.value = ""; c.fill = fill(C_FILL_INPUT)
            c.font = F(C_FONT_INPUT, 10); c.alignment = WRAP; c.border = BORDER
            c = ws.cell(row, 6); c.value = note; c.fill = fill(bg)
            c.font = F(C_FONT_NOTE, 9); c.alignment = WRAP; c.border = BORDER
            ws.row_dimensions[row].height = 42
            row += 1

    widths = {"A": 5, "B": 24, "C": 10, "D": 52, "E": 52, "F": 44}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"
    return ws


# ===========================================================================
# 子表2：监控问题与应答逻辑确认（已删除「应答逻辑备注」列）
# ===========================================================================
def read_response_logic(path):
    """读取企业已填写的《应答逻辑确认表》。
    返回 [(kind, 问题, 应答逻辑, 参考资料)]；kind=='SECTION' 表示区段标题行。
    注意：原表第 3 列为「应答逻辑备注」，按用户要求不再读取；参考资料取原表第 4 列。
    """
    rows = []
    if not path or not os.path.exists(path):
        return rows
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = None
    for r in range(1, min(12, ws.max_row + 1)):
        vals = [str(ws.cell(r, c).value or "") for c in range(1, 5)]
        if any(v.strip() == "问题" for v in vals) and any("应答逻辑" in v for v in vals):
            header_row = r
            break
    if header_row is None:
        header_row = 4
    for r in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value  # 原「应答逻辑备注」——按要求丢弃
        d = ws.cell(r, 4).value  # 原「参考资料」
        if a is None and b is None and c is None and d is None:
            continue
        a_str = str(a).strip() if a is not None else ""
        if a_str and (b is None or str(b).strip() == ""):
            rows.append(("SECTION", a_str, "", ""))
        else:
            rows.append((None, a_str, str(b or ""), str(d or "")))
    return rows


def build_sheet2(wb, brand, response_rows, template_mode):
    ws = wb.create_sheet("监控问题与应答逻辑确认")
    headers = ["#", "监控问题", "应答逻辑（预填）", "✏️ 企业想修改", "参考资料"]
    ncols = len(headers)
    write_title_block(
        ws, ncols,
        f"{brand}  监控问题与应答逻辑确认表",
        "FrontMind 品牌优化工作流  ·  S10 监控问题与应答逻辑确认   "
        "请在黄色「企业想修改」列填写调整意见，留空表示确认应答逻辑无误",
    )
    write_header_row(ws, 5, headers)

    row = 6
    alt = False
    idx = 0

    if template_mode or not response_rows:
        # 模板模式：给出空白示意区段与若干空行，不预填任何企业问题
        write_section_row(ws, row, ncols, "一、监控问题与应答逻辑（运行时由已填写的《应答逻辑确认表》逐条承接）")
        row += 1
        for _ in range(8):
            idx += 1
            bg = C_ROW_ALT if alt else C_WHITE
            alt = not alt
            c = ws.cell(row, 1); c.value = idx; c.fill = fill(bg)
            c.font = F(C_FONT_IDX, 9); c.alignment = CENTER; c.border = BORDER
            for col in (2, 3, 5):
                c = ws.cell(row, col); c.value = ""; c.fill = fill(bg)
                c.font = F(C_FONT_PREFILL, 10); c.alignment = WRAP; c.border = BORDER
            c = ws.cell(row, 4); c.value = ""; c.fill = fill(C_FILL_INPUT)
            c.font = F(C_FONT_INPUT, 10); c.alignment = WRAP; c.border = BORDER
            ws.row_dimensions[row].height = 50
            row += 1
    else:
        has_section = any(r[0] == "SECTION" for r in response_rows)
        if not has_section:
            write_section_row(ws, row, ncols, "一、监控问题与应答逻辑（来自应答逻辑确认表）")
            row += 1
        for kind, q, logic, ref in response_rows:
            if kind == "SECTION":
                write_section_row(ws, row, ncols, q)
                row += 1
                continue
            idx += 1
            bg = C_ROW_ALT if alt else C_WHITE
            alt = not alt
            c = ws.cell(row, 1); c.value = idx; c.fill = fill(bg)
            c.font = F(C_FONT_IDX, 9); c.alignment = CENTER; c.border = BORDER
            c = ws.cell(row, 2); c.value = q; c.fill = fill(bg)
            c.font = F(C_FONT_ITEM, 10, True); c.alignment = WRAP; c.border = BORDER
            c = ws.cell(row, 3); c.value = logic or "（待补充）"; c.fill = fill(bg)
            c.font = F(C_FONT_PREFILL, 10); c.alignment = WRAP; c.border = BORDER
            c = ws.cell(row, 4); c.value = ""; c.fill = fill(C_FILL_INPUT)
            c.font = F(C_FONT_INPUT, 10); c.alignment = WRAP; c.border = BORDER
            c = ws.cell(row, 5); c.value = ref; c.fill = fill(bg)
            c.font = F(C_FONT_NOTE, 9); c.alignment = WRAP; c.border = BORDER
            ws.row_dimensions[row].height = 56
            row += 1

    widths = {"A": 5, "B": 38, "C": 58, "D": 38, "E": 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"
    return ws


# ===========================================================================
def main():
    ap = argparse.ArgumentParser(description="S10 品牌信息确认表生成器")
    ap.add_argument("--brand", default=None, help="品牌名；模板模式可省略")
    ap.add_argument("--mode", choices=["template", "fill"], default="template",
                    help="template=纯空白模板（默认）；fill=整合 S1-S9 与应答逻辑表预填")
    ap.add_argument("--work-dir", default=None, help="策略层成果工作目录（fill 模式必填）")
    ap.add_argument("--response-logic", default=None,
                    help="企业已填写的《应答逻辑确认表》xlsx 路径（fill 模式必填）")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    mode = args.mode
    template_mode = (mode == "template")
    brand = args.brand or ("【品牌名】" if template_mode else "品牌")

    if template_mode:
        out = args.out or "品牌信息确认表_模板.xlsx"
        prefill_map = {}
        response_rows = []
        print("[info] 模板模式：生成纯空白模板，不预填任何企业信息")
    else:
        if not args.work_dir or not args.response_logic:
            print("[error] fill 模式需同时提供 --work-dir 与 --response-logic", file=sys.stderr)
            sys.exit(2)
        wd = args.work_dir
        out = args.out or os.path.join(wd, f"S10_{brand}_品牌信息确认表.xlsx")
        s1 = load_json(find_file(wd, f"S1_{brand}_*.json", "S1_*品牌事实*.json", "S1_*.json"))
        s4 = load_json(find_file(wd, f"S4_{brand}_*.json", "S4_*定位*.json", "S4_*.json"))
        s6 = load_json(find_file(wd, f"S6_{brand}_*.json", "S6_*token*.json", "S6_*.json"))
        prefill_map = build_prefill_map(s1, s4, s6)
        response_rows = read_response_logic(args.response_logic)
        n = sum(1 for r in response_rows if r[0] is None)
        print(f"[info] fill 模式：应答逻辑确认表解析到 {n} 条问题；子表1 预填命中 {len(prefill_map)} 项")

    wb = openpyxl.Workbook()
    build_sheet1(wb, brand, prefill_map, template_mode)
    build_sheet2(wb, brand, response_rows, template_mode)
    wb.save(out)
    print(f"[ok] 已生成品牌信息确认表：{out}")
    return out


if __name__ == "__main__":
    main()
