#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
S10 品牌信息确认表 — 回填差异解析器 (Confirmation Diff Parser)
==============================================================
用途：在「暂停4：品牌信息确认表最终确认」之后，解析企业已回填的
《S10_{brand}_品牌信息确认表.xlsx》，提取企业在两个子表中间高亮列里
填写的全部修改意见，输出一份结构化的《品牌信息修改清单》JSON，供 S0
按方案 A（逐项重新调用受影响的 S 节点）回灌到 S1-S9 源文件。

设计原则：
  - 只把「企业填写 / 修改」「企业想修改」列中**非空**的单元格视为修改项；留空 = 确认无误。
  - 不直接改写任何 S 源文件（回灌由 S0 调度对应 S 节点完成），本脚本只产出清单。
  - 为每条修改项标注其建议影响的 S 节点（target_nodes），便于 S0 路由。

输入：
  --confirmed   企业已回填的《品牌信息确认表》xlsx 路径（必填）
  --brand       品牌名（用于输出文件名，可选）
  --out         输出修改清单 JSON 路径（可选，默认 {brand}_品牌信息修改清单.json）

输出 JSON 结构：
  {
    "brand": "...",
    "source_file": "...",
    "has_changes": true/false,
    "summary": {"sheet1_changes": N, "sheet2_changes": M, "total": N+M,
                "affected_nodes": ["S1","S4",...]},
    "sheet1_changes": [
        {"index": 3, "section": "一、品牌基础信息", "item": "企业全称",
         "prefilled": "...", "requested_change": "...", "target_nodes": ["S1"]} , ...],
    "sheet2_changes": [
        {"index": 5, "section": "...", "question": "...",
         "prefilled_logic": "...", "requested_change": "...",
         "target_nodes": ["S5","S8","应答逻辑确认表"]}, ...]
  }
"""

import argparse
import json
import os
import sys

import openpyxl

# ---------------------------------------------------------------------------
# 子表1 条目 → 受影响 S 节点 的映射（与生成器 SHEET1_SECTIONS 对应）
# 用「区段」给出默认归属，必要时按条目名称细化。
# ---------------------------------------------------------------------------
SECTION_TO_NODES = {
    "一、品牌基础信息": ["S1"],
    "二、品牌定位与价值主张": ["S4"],
    "三、产品 / 服务矩阵": ["S1", "S2"],
    "四、核心优势": ["S4", "S1"],
    "五、目标客群": ["S2"],
    "六、资质认证": ["S1"],
    "七、客户案例": ["S1", "S7"],
    "八、量化经营数据": ["S1"],
    "九、竞争格局": ["S3", "S5"],
    "十、服务区域": ["S1"],
    "十一、品牌历史与里程碑": ["S1"],
    "十二、备注与想法": ["S6"],
    "十三、补充材料清单": ["S1"],
}

# 子表2（监控问题与应答逻辑）默认受影响节点
SHEET2_NODES = ["S5", "S8", "应答逻辑确认表"]

EMPTY_PLACEHOLDERS = {
    "", "（由策略层 S1-S9 自动填充）", "（待补充）", "无", "n/a", "na", "none", "-",
}


def _norm(v):
    return str(v).strip() if v is not None else ""


def _is_empty(v):
    return _norm(v).lower() in EMPTY_PLACEHOLDERS


def _find_header_row(ws, must_have):
    """在前若干行里定位表头行。
    must_have 是「关键词组」的列表，例如 [["条目名称","条目"], ["企业填写","修改"]]；
    某行被认定为表头，当且仅当每个关键词组内至少有一个关键词出现在该行某列。"""
    for r in range(1, min(12, ws.max_row + 1)):
        vals = [_norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if all(any(kw in v for v in vals for kw in group) for group in must_have):
            return r, vals
    return None, None


def _col_index(header_vals, *keywords):
    """在表头行里定位列号（1-based）。
    优先精确包含靠前的关键词；为避免“问题”误中“企业想修改”等，
    按 keywords 顺序逐个尝试，同一关键词在所有列中找最靠前的命中。找不到返回 None。"""
    for kw in keywords:
        for i, v in enumerate(header_vals, start=1):
            if kw in v:
                return i
    return None


def parse_sheet1(ws):
    """解析子表1，返回非空修改项列表。"""
    changes = []
    header_row, header_vals = _find_header_row(ws, [["#"], ["条目名称"], ["企业填写"]])
    if header_row is None:
        header_row, header_vals = 5, [_norm(ws.cell(5, c).value) for c in range(1, ws.max_column + 1)]
    col_item = _col_index(header_vals, "条目名称", "条目") or 2
    col_prefill = _col_index(header_vals, "已有内容", "预填") or 4
    col_input = _col_index(header_vals, "企业填写", "修改") or 5

    current_section = ""
    idx = 0
    for r in range(header_row + 1, ws.max_row + 1):
        first = _norm(ws.cell(r, 1).value)
        item = _norm(ws.cell(r, col_item).value)
        # 区段标题行：第一列为空或非数字，且条目列空 —— 用合并区段判断
        if item == "" and first == "":
            continue
        if item == "" and first and not first.isdigit():
            current_section = first
            continue
        # 数据行
        if first.isdigit():
            idx = int(first)
        prefilled = _norm(ws.cell(r, col_prefill).value)
        requested = _norm(ws.cell(r, col_input).value)
        if not _is_empty(requested):
            changes.append({
                "index": idx,
                "section": current_section,
                "item": item,
                "prefilled": "" if _is_empty(prefilled) else prefilled,
                "requested_change": requested,
                "target_nodes": SECTION_TO_NODES.get(current_section, ["S1"]),
            })
    return changes


def parse_sheet2(ws):
    """解析子表2（监控问题与应答逻辑），返回非空修改项列表。"""
    changes = []
    header_row, header_vals = _find_header_row(ws, [["#"], ["监控问题"], ["企业想修改"]])
    if header_row is None:
        header_row, header_vals = 5, [_norm(ws.cell(5, c).value) for c in range(1, ws.max_column + 1)]
    col_q = _col_index(header_vals, "监控问题") or 2
    col_logic = _col_index(header_vals, "应答逻辑") or 3
    col_input = _col_index(header_vals, "企业想修改") or 4

    current_section = ""
    idx = 0
    for r in range(header_row + 1, ws.max_row + 1):
        first = _norm(ws.cell(r, 1).value)
        q = _norm(ws.cell(r, col_q).value)
        if q == "" and first == "":
            continue
        if q == "" and first and not first.isdigit():
            current_section = first
            continue
        if first.isdigit():
            idx = int(first)
        logic = _norm(ws.cell(r, col_logic).value)
        requested = _norm(ws.cell(r, col_input).value)
        if not _is_empty(requested):
            changes.append({
                "index": idx,
                "section": current_section,
                "question": q,
                "prefilled_logic": "" if _is_empty(logic) else logic,
                "requested_change": requested,
                "target_nodes": SHEET2_NODES,
            })
    return changes


def parse_confirmation(path, brand=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    # 子表1：优先按名称匹配，否则取第一个 sheet
    ws1 = wb["品牌信息确认"] if "品牌信息确认" in wb.sheetnames else wb[wb.sheetnames[0]]
    ws2 = None
    for name in wb.sheetnames:
        if "应答逻辑" in name or "监控问题" in name:
            ws2 = wb[name]
            break
    sheet1_changes = parse_sheet1(ws1)
    sheet2_changes = parse_sheet2(ws2) if ws2 is not None else []

    affected = []
    for ch in sheet1_changes + sheet2_changes:
        for n in ch["target_nodes"]:
            if n not in affected:
                affected.append(n)

    result = {
        "brand": brand or "",
        "source_file": os.path.basename(path),
        "has_changes": bool(sheet1_changes or sheet2_changes),
        "affected_nodes": affected,  # 顶层镜像，便于 S0 与下游脚本直接读取
        "summary": {
            "sheet1_changes": len(sheet1_changes),
            "sheet2_changes": len(sheet2_changes),
            "total": len(sheet1_changes) + len(sheet2_changes),
            "affected_nodes": affected,
        },
        "sheet1_changes": sheet1_changes,
        "sheet2_changes": sheet2_changes,
    }
    return result


def main():
    ap = argparse.ArgumentParser(description="S10 品牌信息确认表回填差异解析器")
    ap.add_argument("--confirmed", required=True,
                    help="企业已回填的《品牌信息确认表》xlsx 路径")
    ap.add_argument("--brand", default=None, help="品牌名（用于输出文件名）")
    ap.add_argument("--out", default=None, help="输出修改清单 JSON 路径")
    args = ap.parse_args()

    if not os.path.exists(args.confirmed):
        print(f"[error] 找不到确认表文件：{args.confirmed}", file=sys.stderr)
        sys.exit(2)

    result = parse_confirmation(args.confirmed, args.brand)
    brand = args.brand or result.get("brand") or "品牌"
    out = args.out or f"{brand}_品牌信息修改清单.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    s = result["summary"]
    if result["has_changes"]:
        print(f"[ok] 检出修改项：子表1 {s['sheet1_changes']} 条 / 子表2 {s['sheet2_changes']} 条，"
              f"涉及节点 {s['affected_nodes']}")
        print(f"[ok] 修改清单已写出：{out}")
        print("[next] S0 应按 affected_nodes 逐项重新调用对应 S 节点回灌，再重新封装策略包。")
    else:
        print("[ok] 企业未填写任何修改意见，视为全部确认无误。")
        print(f"[ok] 空清单已写出：{out}")
        print("[next] S0 可直接进入最终策略包封装。")
    return out


if __name__ == "__main__":
    main()
